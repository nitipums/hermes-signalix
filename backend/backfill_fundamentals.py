"""Backfill company profile fundamentals from SET financial-statement workbooks.

The workbook formats differ by issuer.  This module deliberately discovers labels
rather than relying on fixed sheet names or column positions.  Optional fields
remain NULL when the workbook does not state them unambiguously.
"""
from __future__ import annotations

import argparse
import json
import os
import re
import zipfile
from pathlib import Path
from typing import Any

import openpyxl
import psycopg2

FIN_DIR = Path(os.getenv("SET50_FINANCIALS_DIR", "/root/set50_financials"))
DSN = {
    "host": os.getenv("POSTGRES_HOST", "127.0.0.1"),
    "port": os.getenv("POSTGRES_PORT", "5432"),
    "user": os.getenv("POSTGRES_USER", "signalix"),
    "password": os.getenv("POSTGRES_PASSWORD", "signalix_pass"),
    "dbname": os.getenv("POSTGRES_DB", "signalix"),
}
_PAID_LABEL = re.compile(r"ทุนที่ออก\s*(?:และ)?\s*ชำระ|issued\s+and\s+(?:fully\s+)?paid", re.I)
_SHARE_TEXT = re.compile(r"(?:หุ้นสามัญ|ordinary\s+shares?)\s*(?:จำนวน\s*)?[:：]?\s*([0-9][0-9,]*(?:\.\d+)?)\s*(ล้าน|พัน)?", re.I)
_PERCENT = re.compile(r"(?<![\d.])([0-9]{1,3}(?:\.\d+)?)\s*%")


def _text(value: Any) -> str:
    return "" if value is None else str(value).replace("\n", " ").strip()


def _share_count(text: str) -> int | None:
    match = _SHARE_TEXT.search(text)
    if not match:
        return None
    try:
        value = float(match.group(1).replace(",", ""))
    except ValueError:
        return None
    unit = match.group(2)
    if unit == "ล้าน":
        value *= 1_000_000
    elif unit == "พัน":
        value *= 1_000
    return int(value) if value >= 1000 and value.is_integer() else None


def _explicit_percent(text: str, labels: tuple[str, ...]) -> float | None:
    lowered = text.lower()
    if not any(label.lower() in lowered for label in labels):
        return None
    match = _PERCENT.search(text)
    if not match:
        return None
    value = float(match.group(1))
    return value if 0 <= value <= 100 else None


def extract_shares_outstanding(workbook: Any) -> int | None:
    """Find paid-up ordinary shares, independent of sheet/column naming.

    Only a share-count row associated with an issued/paid label is accepted. This
    avoids mistaking authorised shares or weighted-average shares for outstanding
    shares.
    """
    candidates: list[tuple[int, int, int]] = []
    for ws in workbook.worksheets:
        rows = [[_text(v) for v in row] for row in ws.iter_rows(values_only=True)]
        for index, row in enumerate(rows):
            row_text = " | ".join(v for v in row if v)
            shares = _share_count(row_text)
            if shares is None:
                continue
            score = 0
            if _PAID_LABEL.search(row_text):
                score += 10
            for previous in rows[max(0, index - 3):index]:
                if _PAID_LABEL.search(" | ".join(v for v in previous if v)):
                    score += 8
                    break
            if score:
                candidates.append((score, -index, shares))
    if not candidates:
        return None
    candidates.sort(reverse=True)
    return candidates[0][2]


def extract_optional_percentages(workbook: Any) -> tuple[float | None, float | None]:
    """Read only explicitly labelled percentages; otherwise return NULLs."""
    free_float = None
    foreign_limit = None
    for ws in workbook.worksheets:
        for row in ws.iter_rows(values_only=True):
            text = " | ".join(_text(v) for v in row if _text(v))
            free_float = free_float if free_float is not None else _explicit_percent(
                text, ("free float", "สัดส่วนผู้ถือหุ้นรายย่อย")
            )
            foreign_limit = foreign_limit if foreign_limit is not None else _explicit_percent(
                text, ("foreign ownership limit", "foreign limit", "สัดส่วนการถือหุ้นของคนต่างด้าว")
            )
    return free_float, foreign_limit


def find_year_workbook(symbol: str, fin_dir: Path = FIN_DIR) -> Path | None:
    directory = fin_dir / symbol / "year2568"
    if not directory.is_dir():
        return None
    for path in directory.iterdir():
        if path.is_file() and path.name.lower() == "financial_statements.xlsx":
            return path
    return None


def process_symbol(symbol: str, latest_close: float | None, fin_dir: Path = FIN_DIR) -> dict[str, Any] | None:
    path = find_year_workbook(symbol, fin_dir)
    if path is None:
        return None
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    shares = extract_shares_outstanding(workbook)
    free_float, foreign_limit = extract_optional_percentages(workbook)
    market_cap = int(round(shares * latest_close)) if shares is not None and latest_close is not None else None
    return {
        "symbol": symbol,
        "shares_outstanding": shares,
        "market_cap": market_cap,
        "free_float_pct": free_float,
        "foreign_limit_pct": foreign_limit,
        "file": str(path),
    }


def _latest_ord_closes(cur: Any) -> dict[str, float]:
    cur.execute("""SELECT DISTINCT ON (symbol) symbol, close
                   FROM price_data
                   WHERE instrument_type = 'ORD' AND close IS NOT NULL
                   ORDER BY symbol, date DESC""")
    return {symbol: float(close) for symbol, close in cur.fetchall()}


def upsert_profile(cur: Any, data: dict[str, Any]) -> None:
    cur.execute(
        """INSERT INTO company_profiles
          (symbol, source, fetched_at, shares_outstanding, market_cap, free_float_pct, foreign_limit_pct)
          VALUES (%s, 'xlsx', NOW(), %s, %s, %s, %s)
          ON CONFLICT (symbol) DO UPDATE SET
            shares_outstanding = COALESCE(EXCLUDED.shares_outstanding, company_profiles.shares_outstanding),
            market_cap = COALESCE(EXCLUDED.market_cap, company_profiles.market_cap),
            free_float_pct = COALESCE(EXCLUDED.free_float_pct, company_profiles.free_float_pct),
            foreign_limit_pct = COALESCE(EXCLUDED.foreign_limit_pct, company_profiles.foreign_limit_pct),
            source = CASE WHEN EXCLUDED.shares_outstanding IS NOT NULL
                               OR EXCLUDED.market_cap IS NOT NULL
                               OR EXCLUDED.free_float_pct IS NOT NULL
                               OR EXCLUDED.foreign_limit_pct IS NOT NULL
                          THEN 'xlsx' ELSE company_profiles.source END,
            fetched_at = NOW()""",
        (data["symbol"], data["shares_outstanding"], data["market_cap"], data["free_float_pct"], data["foreign_limit_pct"]),
    )


def run(fin_dir: Path = FIN_DIR, pg: Any | None = None) -> dict[str, Any]:
    owns_connection = pg is None
    pg = pg or psycopg2.connect(**DSN)
    cur = pg.cursor()
    try:
        closes = _latest_ord_closes(cur)
        symbols = sorted(p.name.upper() for p in fin_dir.iterdir() if p.is_dir())
        report = {"files": 0, "symbols": len(symbols), "parsed_shares": 0, "free_float": 0, "foreign_limit": 0, "rows_upserted": 0, "missing_files": [], "errors": [], "details": []}
        for symbol in symbols:
            try:
                data = process_symbol(symbol, closes.get(symbol), fin_dir)
            except (OSError, ValueError, zipfile.BadZipFile) as exc:
                report["errors"].append({"symbol": symbol, "error": str(exc)})
                continue
            if data is None:
                report["missing_files"].append(symbol)
                continue
            report["files"] += 1
            report["parsed_shares"] += data["shares_outstanding"] is not None
            report["free_float"] += data["free_float_pct"] is not None
            report["foreign_limit"] += data["foreign_limit_pct"] is not None
            upsert_profile(cur, data)
            report["rows_upserted"] += 1
            report["details"].append({k: v for k, v in data.items() if k != "file"})
        pg.commit()
        return report
    finally:
        cur.close()
        if owns_connection:
            pg.close()


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--fin-dir", type=Path, default=FIN_DIR)
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args()
    report = run(args.fin_dir)
    if args.json:
        print(json.dumps(report, ensure_ascii=False, indent=2))
    else:
        print(json.dumps({k: v for k, v in report.items() if k != "details"}, ensure_ascii=False))
        for detail in report["details"]:
            print(f"{detail['symbol']}: shares={detail['shares_outstanding']} market_cap={detail['market_cap']} free_float={detail['free_float_pct']} foreign_limit={detail['foreign_limit_pct']}")


if __name__ == "__main__":
    main()
